from cProfile import label
import numpy as np
import pandas as pd
import os
import statsmodels.api as sm
import sys
import openpyxl
import matplotlib as mpl
import matplotlib.pylab
import matplotlib.pyplot as plt
import math

def get_value(column0):
    list0 = []
    for i in column0:
        list0.append(i.value)
    list0.pop(0)
    if not list0[-1]:
        list0.pop(-1)
    return list0

'''设置目录并打开Excel'''
ml1 = os.getcwd()
os.chdir('./供需比的四带计算')
ws = openpyxl.load_workbook('4ring - py.xlsx')['4ring']

'''索引第一行title_name，按name获取列值'''
title = {}
value = {}
for i, d in enumerate(ws[1]):
    letter = openpyxl.utils.get_column_letter(i+1)
    name = d.value
    title[letter] = name
    value[name] = get_value(ws[letter])

'''标准比较
r1/r10, [0.1,0.2], r1_N_SD, r10_N_SD
r2, [0.83,2], r2_N_SD
'''

result_r = {}
for i in range(4):
    result_r[i+1] = {'below':0,'std':0,'above':0}

r1 = value['r1_N_SD']
r10 = value['r10_N_SD']
r2 = value['r2_N_SD']

lim = [0.833, 2]
r = r2

for i0, v0 in enumerate( value['group_4']):
    if r[i0] < lim[0]:
        result_r[v0]['below'] += 1
    if lim[0] <= r[i0] < lim[1]:
        result_r[v0]['std'] += 1
    if r[i0] > lim[1]:
        result_r[v0]['above'] += 1
outwb = openpyxl.load_workbook('out_result.xlsx')
outws = outwb['1']
cell_range = outws['A1':'C12']
out_data = []
for k0, v0 in result_r.items():
    for k1, v1 in v0.items():
        out_data.append(k0)
        out_data.append(k1)
        out_data.append(v1)
out_data
for i0, line in enumerate(cell_range):
    for i1, cell in enumerate(line):
        cell.value = out_data[i0*3+i1]
outwb.save('out_result.xlsx')
outwb.close()
print('...succeed')


    
    
    


