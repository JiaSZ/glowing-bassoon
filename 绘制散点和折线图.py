from cProfile import label
import numpy as np
import pandas as pd
import os
import statsmodels.api as sm
import sys
import openpyxl
import matplotlib as mpl
import matplotlib.pylab
import matplotlib.pyplot as plt
import math

def get_value(column0):
    list0 = []
    for i in column0:
        list0.append(i.value)
    list0.pop(0)
    return list0

def step_x(x, n):
    x = value[x]
    x_km = []
    range_x = math.ceil( max(x)/n )
    for i in range(range_x):
        x_km.append( n/2 + i * n)
    return x_km

def aver_y(y, x):
    y_aver = []
    for i in range( x_range ):
        x_index = [i2 for i2,v in enumerate(value_x) if (i*n <= v < (i+1)*n)]
        y_aver_c = np.average([y[i2] for i2 in x_index])
        y_aver.append(y_aver_c if math.isnan(y_aver_c) == False else \
            np.average(y_aver[-3:])  )
    return y_aver
    
def sum_p(x, p):
    p_sum = []
    for i in range( x_range ):
        x_index = [i2 for i2,v in enumerate(value_x) if (i*n <= v < (i+1)*n)]
        p_sum_c = [p[i2] for i2 in x_index]
        p_sum.append( sum(p_sum_c)/10000 if sum(p_sum_c)/10000 else 0) 
    return p_sum

def sum_c(x, c):
    c_sum = []
    for i in range( x_range ):
        x_index = [i2 for i2,v in enumerate(value_x) if (i*n <= v < (i+1)*n)]
        c_sum_c = [c[i2] for i2 in x_index]
        c_sum.append( sum(c_sum_c) )
    return c_sum

def scatter0(y):
    Y = value[y]
    X = value_x
    fig = plt.figure(figsize=(8, 6), dpi = 100)
    # plt.style.use('seaborn')

    plt.scatter(X, Y, s = 9, label = y)
    plt.xlabel(str(x) + ' (km)')
    plt.ylabel(y)
    plt.xlim(0,120)
    # plt.ylim(-0.1,3)
    
    '''趋势线'''
    # z = np.polyfit(X, Y, 1)
    # p = np.poly1d(z)
    # zlist0 = []
    # for i in z:
    #     temp = '%.2e' % i
    #     zlist0.append(temp)
    # label2 = 'Y = a2 * X^2 + a1 * X + b\n' + str(zlist0)
    # mpl.pylab.plot(X,p(X),"r--", label = label2, c = 'r')

    ''' n 公里平均值,[)'''

    y_aver = aver_y(Y, X)

    plt.plot(x_list, y_aver, label = 'aver_by_' + str(n) + 'km', \
        marker = '+' ,c = 'r', linestyle="-", linewidth=1.5)

    plt.legend(loc = 'upper left')
    print(plt.style.available)
    
    plt.show()

def scatter0_add_p(y, p):
    
    Y = value[y]
    X = value_x
    P = value[p]
    fig, ax1 = plt.subplots(figsize=(8, 6), dpi = 120)
    ax2 = ax1.twinx()
    ax1.scatter(X, Y, s = 9, label = y)
    ax1.set_xlabel(str(x) + ' (km)')
    ax1.set_ylabel(y)
    ax1.set_xlim(0, 120)

    '''n 公里总人口[)'''

    p_sum = sum_p(X, P)
    ax2.plot(x_list, p_sum, label = 'p_sum (*10,000)', c = 'b',\
        marker = '+', mec = 'k')
    ax2.set_ylim(0,450)
    y_aver = aver_y(Y, X)
    ax1.plot(x_list, y_aver, label = 'aver_by_' + str(n) + 'km', \
        marker = '+', mec='k', c = 'r', linestyle="-", linewidth=1.5)
    ax1.legend(loc = 'upper left')
    ax2.legend(loc = 'upper right')
    plt.show()



def scatter0_p_and_c(c1, c2, p):
    c1 = value[c1]
    c2 = value[c2]
    x = value_x
    p = value[p]

    fig, ax1 = plt.subplots(figsize=(8, 6), dpi = 120)
    ax2 = ax1.twinx()

    c1 = sum_c(x, c1)
    c2 = sum_c(x, c2)

    ax1.set_xlim(0, 120)
    # ax1.set_ylim(0, 300)
    ax1.plot(x_list, c1, label = 'C_r1')
    ax1.plot(x_list, c2, label = 'C_r2')

    p_sum = sum_p(x, p)
    ax2.plot(x_list, p_sum, label = 'p_sum (*10,000)', c = 'b',\
        marker = '+', mec = 'k')
    # ax2.set_ylim(0, 500)
    ax1.legend(loc = 'upper left')
    ax2.legend(loc = 'upper right')
    plt.show()
    


if __name__ == '__main__':

    '''设置目录并打开Excel'''
    ml1 = os.getcwd()
    os.chdir('./距离计算')
    ws = openpyxl.load_workbook('Dist.xlsx')['Dist']

    '''索引第一行title_name，按name获取列值'''
    title = {}
    value = {}
    for i, d in enumerate(ws[1]):
        letter = openpyxl.utils.get_column_letter(i+1)
        name = d.value
        title[letter] = name
        value[name] = get_value(ws[letter])

    '''绘图'''
    y = 'r12_N_SD'
    x = 'NEAR_DIST'
    p = 'P_2020E'

    n = 5
    value_x = [i/1000 for i in value[x]]
    x_list = step_x(x, n)
    x_range = len(x_list)


    scatter0(y)
    scatter0_add_p(y, p)

    c1 = 'C_r1'
    c2 = 'C_r2'
    scatter0_p_and_c(c1, c2, p)

    



